from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.db import transaction
from django.db.models import Max

from django.forms import modelform_factory
from django.shortcuts import render, redirect, get_object_or_404
from .models import Region
from .forms import RegionForm
from django.db.models import Q
from django.urls import reverse
from urllib.parse import urlencode
from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse
from django.db.models import Q
from .forms import BatchForm
from .forms import ReceivingNormalSKRForm
from .forms import ReceivingSKRWithMoistureForm
from .forms import RegionForm, StoreForm, FieldOfficerForm
from django.db.models import Q
from datetime import datetime
from django.core.paginator import Paginator

#  ************** BATCHES  *************
from django.db import IntegrityError
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.timezone import now
from .forms import BatchForm
from .models import Batch, ReceivingNormalSKR, ReceivingSKRWithMoisture
from django.db.models import Q
from datetime import datetime
from django.core.paginator import Paginator

@login_required
def create_batch(request):
    last_lot = None

    if request.method == 'POST':
        form = BatchForm(request.POST)
        if form.is_valid():
            try:
                batch = form.save(commit=False)
                batch.created_by = request.user
                batch.save()

                print(f"Batch created successfully: {batch.batch_number}")
                return redirect('receiving_detail', pk=batch.pk)
            except IntegrityError:
                form.add_error(None, "A batch with the generated number already exists. Please try again with "
                                     "different inputs or contact support if the issue persists.")
            except Exception as e:
                form.add_error(None, f"An unexpected error occurred: {e}")
        else:
            print("Form is NOT valid. Errors:")
            print(form.errors)

    else:
        form = BatchForm()
        current_year = now().year
        latest = Batch.objects.filter(created_at__year=current_year).order_by('-lot_number').first()
        if latest:
            last_lot = latest.lot_number

    return render(request, 'inventory/create_batch.html', {'form': form, 'last_lot': last_lot})

@login_required
def update_batch(request, pk):
    """
    Handles the updating of an existing Batch instance.
    """
    batch = get_object_or_404(Batch, pk=pk)
    if request.method == 'POST':
        form = BatchForm(request.POST, instance=batch)
        if form.is_valid():
            if form.has_changed():  # Check if any data actually changed
                try:
                    form.save()
                    return redirect('receiving_detail', pk=batch.pk)
                except IntegrityError:
                    # Catch database-level unique constraint violations for the regenerated batch_number
                    form.add_error(None,
                                   "The updated batch details would result in a duplicate batch number. Please change "
                                   "the certification, lot number, or drier number.")
                except Exception as e:
                    form.add_error(None, f"An unexpected error occurred during update: {e}")
            else:
                form.add_error(None, "No changes detected. Please modify the fields to update.")
    else:
        form = BatchForm(instance=batch)

    return render(request, 'inventory/update_batch.html', {'form': form, 'batch': batch})

@login_required
def delete_batch(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    if request.method == 'POST':
        batch.delete()
        return redirect('batch_list')
    return render(request, 'inventory/delete_batch.html', {'batch': batch})

@login_required
def batch_list(request):
    batches = Batch.objects.all().order_by('-created_at')

    query = request.GET.get('q')
    cert_filter = request.GET.get('certification')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if query:
        batches = batches.filter(
            Q(batch_number__icontains=query) |
            Q(vehicle_registration__icontains=query) |
            Q(driver_name__icontains=query) |
            Q(lot_number__icontains=query) |
            Q(receivingnormalskr__field_officer__store__name__icontains=query) |
            Q(receivingnormalskr__field_officer__store__region__name__icontains=query)
        ).distinct()

    if cert_filter:
        batches = batches.filter(certification__iexact=cert_filter)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d")
            batches = batches.filter(created_at__date__gte=date_from_obj)
        except:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d")
            batches = batches.filter(created_at__date__lte=date_to_obj)
        except:
            pass

    # Pagination
    paginator = Paginator(batches, 20)
    page_number = request.GET.get('page')
    batches = paginator.get_page(page_number)

    context = {
        'batches': batches,
        'query': query or '',
        'cert_filter': cert_filter or '',
        'date_from': date_from or '',
        'date_to': date_to or '',
    }
    return render(request, 'inventory/batch_list.html', context)


@login_required
def batch_detail(request, pk):
    batch = get_object_or_404(Batch, pk=pk)

    skr_normal_samples = ReceivingNormalSKR.objects.filter(batch=batch)
    skr_moisture_samples = ReceivingSKRWithMoisture.objects.filter(batch=batch)

    region_set = set()
    officer_set = set()

    for sample in list(skr_normal_samples) + list(skr_moisture_samples):
        fo = sample.field_officer
        if fo:
            # Get region from FO > Store > Region
            if fo.store and fo.store.region:
                region_set.add(fo.store.region.name)

            # Add officer name
            officer_set.add(fo.name)

    context = {
        'batch': batch,
        'regions_joined': ' & '.join(sorted(region_set)) if region_set else None,
        'officers_joined': ' & '.join(sorted(officer_set)) if officer_set else None,
        'skr_normal_samples': skr_normal_samples,
        'skr_moisture_samples': skr_moisture_samples,
    }
    return render(request, 'inventory/batch_detail.html', context)

# ******************************************** ReceivingNormalSKR ***************************
from django.contrib import messages
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.db.models import Max
from decimal import Decimal, InvalidOperation
from django.contrib import messages # Don't forget to import messages
from django.urls import reverse # Don't forget to import reverse

# Assuming these are your models
from .models import Batch, ReceivingNormalSKR, Region, Store, FieldOfficer, Sizing


@login_required
def add_normal_skr(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    regions = Region.objects.all()

    # Initialize variables to hold selected values for re-rendering form on error
    selected_region_id = None
    selected_store_id = None
    fo_id = None
    stores_in_region = []
    field_officers_in_store = []
    form_data = request.POST

    if request.method == 'POST':
        fo_id = request.POST.get('fo_id')
        action = request.POST.get('action')

        if not fo_id:
            errors = ["Please select a Field Officer."]
            # Attempt to re-populate regions, stores, and FOs for error display
            if request.POST.get('region'):
                selected_region_id = request.POST.get('region')
                stores_in_region = Store.objects.filter(region_id=selected_region_id)
            if request.POST.get('store'):
                selected_store_id = request.POST.get('store')
                field_officers_in_store = FieldOfficer.objects.filter(store_id=selected_store_id)

            return render(request, 'inventory/add_normal_skr.html', {
                'batch': batch,
                'regions': regions,
                'errors': errors,
                'selected_region_id': selected_region_id,
                'selected_store_id': selected_store_id,
                'fo_id': fo_id,
                'stores_in_region': stores_in_region,
                'field_officers_in_store': field_officers_in_store,
                'form_data': form_data, # Pass all POST data to repopulate sample fields
            })

        # Get the FieldOfficer object AFTER validation
        fo = get_object_or_404(FieldOfficer, id=fo_id)

        good_qs = request.POST.getlist('good_q[]')
        insect_damageds = request.POST.getlist('insect_damaged[]')
        molds = request.POST.getlist('mold[]')
        immatures = request.POST.getlist('immature[]')
        weights = request.POST.getlist('weight[]')
        moistures = request.POST.getlist('moisture[]')
        spillages = request.POST.getlist('spillage[]')

        size_0s = request.POST.getlist('size_0[]')
        size_1ls = request.POST.getlist('size_1l[]')
        size_1ss = request.POST.getlist('size_1s[]')
        size_1xss = request.POST.getlist('size_1xs[]')

        # Determine the starting sample number for this submission
        last_sample = ReceivingNormalSKR.objects.filter(
            batch=batch, field_officer=fo
        ).aggregate(Max('sample_number'))['sample_number__max'] or 0

        errors = [] # Changed from invalid_rows to a general errors list for consistency
        samples_to_save = [] # To store successfully parsed data

        for idx in range(len(good_qs)):
            try:
                # --- Robust Decimal Conversion ---
                # Use .strip() to remove leading/trailing whitespace
                # Default to '0' if the string is empty after stripping
                good_q = Decimal(good_qs[idx].strip() if good_qs[idx].strip() else '0')
                insect_damaged = Decimal(insect_damageds[idx].strip() if insect_damageds[idx].strip() else '0')
                mold = Decimal(molds[idx].strip() if molds[idx].strip() else '0')
                immature = Decimal(immatures[idx].strip() if immatures[idx].strip() else '0')

                # For fields that might genuinely be empty and map to NULL in DB, use None
                # Assuming 'weight', 'moisture', 'spillage' could be optional in the model (null=True)
                weight_str = weights[idx].strip()
                weight = Decimal(weight_str) if weight_str else None # If empty, store as None

                moisture_str = moistures[idx].strip()
                moisture = Decimal(moisture_str) if moisture_str else None # If empty, store as None

                spillage_str = spillages[idx].strip()
                spillage = Decimal(spillage_str) if spillage_str else Decimal('0') # Default to 0 if empty/missing for spillage

                size_0 = Decimal(size_0s[idx].strip() if size_0s[idx].strip() else '0')
                size_1l = Decimal(size_1ls[idx].strip() if size_1ls[idx].strip() else '0')
                size_1s = Decimal(size_1ss[idx].strip() if size_1ss[idx].strip() else '0')
                size_1xs = Decimal(size_1xss[idx].strip() if size_1xss[idx].strip() else '0')

                total_quality = good_q + insect_damaged + mold + immature
                # Allow for a small floating point tolerance
                if abs(total_quality - 100) > Decimal('0.01'):
                    errors.append(f"Sample #{idx + 1}: Good Quality, Insect Damaged, Mold, and Immature must sum to 100%. Current total: {total_quality}%.")
                    continue # Skip this sample, but continue processing others

                samples_to_save.append({
                    'sample_number': last_sample + idx + 1,
                    'good_q': good_q,
                    'insect_damaged': insect_damaged,
                    'mold': mold,
                    'immature': immature,
                    'weight': weight,
                    'moisture': moisture,
                    'spillage': spillage,
                    'sizing_data': {
                        'size_0': size_0,
                        'size_1l': size_1l,
                        'size_1s': size_1s,
                        'size_1xs': size_1xs,
                    }
                })

            except (ValueError, InvalidOperation):
                errors.append(f"Sample #{idx + 1}: Invalid number format for one or more fields. Please enter valid numbers.")
            except IndexError:
                errors.append(f"Sample #{idx + 1}: Missing data for one or more fields.")

        if errors:
            # Re-populate selected region, store, and FO for display
            # These were captured from request.POST at the top of the function
            selected_region_id = request.POST.get('region')
            selected_store_id = request.POST.get('store')

            if selected_region_id:
                stores_in_region = Store.objects.filter(region_id=selected_region_id)
            if selected_store_id:
                field_officers_in_store = FieldOfficer.objects.filter(store_id=selected_store_id)

            return render(request, 'inventory/add_normal_skr.html', {
                'batch': batch,
                'regions': regions,
                'errors': errors, # Pass the list of errors
                'selected_region_id': selected_region_id,
                'selected_store_id': selected_store_id,
                'fo_id': fo_id,
                'stores_in_region': stores_in_region,
                'field_officers_in_store': field_officers_in_store,
                'form_data': form_data, # Pass all POST data to repopulate sample fields
            })

        # Save only if all samples processed without critical errors
        with transaction.atomic():
            for sample_data in samples_to_save:
                normal_skr_entry = ReceivingNormalSKR.objects.create(
                    batch=batch,
                    field_officer=fo,
                    sample_number=sample_data['sample_number'],
                    good_q=sample_data['good_q'],
                    insect_damaged=sample_data['insect_damaged'],
                    mold=sample_data['mold'],
                    immature=sample_data['immature'],
                    weight=sample_data['weight'],
                    moisture=sample_data['moisture'],
                    spillage=sample_data['spillage']
                )

                Sizing.objects.create(
                    normal_skr_entry=normal_skr_entry,
                    **sample_data['sizing_data'] # Unpack sizing data
                )
            messages.success(request, f"{len(samples_to_save)} SKR sample(s) added successfully for Batch {batch.batch_number}.")

        action = request.POST.get('action')
        if action == 'next':
            return redirect(reverse('add_moisture_skr', kwargs={'pk': batch.pk}))
        else:
            return redirect(reverse('receiving_detail', kwargs={'pk': batch.pk}))

    # Initial GET request
    return render(request, 'inventory/add_normal_skr.html', {
        'batch': batch,
        'regions': regions,
        # No selected IDs or stores/FOs on initial load
        'selected_region_id': None,
        'selected_store_id': None,
        'fo_id': None,
        'stores_in_region': [],
        'field_officers_in_store': [],
    })


@login_required
def get_stores_by_region(request, region_id):
    stores = Store.objects.filter(region_id=region_id).values('id', 'name')
    return JsonResponse(list(stores), safe=False)

@login_required
def get_field_officers_by_store(request, store_id):
    fos = FieldOfficer.objects.filter(store_id=store_id).values('id', 'name')
    return JsonResponse(list(fos), safe=False)

@login_required
def get_store_and_region(request, fo_id):
    try:
        fo = FieldOfficer.objects.select_related('store__region').get(id=fo_id)
        return JsonResponse({
            'store': fo.store.name,
            'region': fo.store.region.name,
        })
    except FieldOfficer.DoesNotExist:
        return JsonResponse({'error': 'Field Officer not found'}, status=404)


# REGION
# ------------------ REGION --------------------
@login_required
def create_region(request):
    # --- Filter Logic Start ---
    search_query = request.GET.get('search')

    regions = Region.objects.all().order_by('name')
    if search_query:
        try:
            search_code = int(search_query)
            regions = regions.filter(Q(name__icontains=search_query) | Q(region_code=search_code))
        except ValueError:
            regions = regions.filter(name__icontains=search_query)
    # --- Filter Logic End ---

    # --- Form Handling for Add/Edit Region ---
    form = RegionForm()
    edit_id = request.GET.get('edit_id')
    edit_form = None

    if edit_id:
        try:
            edit_region_instance = get_object_or_404(Region, id=edit_id)
            edit_form = RegionForm(instance=edit_region_instance)
        except (Region.DoesNotExist, ValueError):
            edit_id = None

    if request.method == 'POST':
        current_get_params = request.GET.dict()

        # --- CONSTRUCT REDIRECT URL WITH QUERY PARAMETERS ---
        # This handles preserving search/edit_id params after POST
        base_redirect_url = reverse('create_region')
        if current_get_params:
            query_string = urlencode(current_get_params)
            final_redirect_url = f"{base_redirect_url}?{query_string}"
        else:
            final_redirect_url = base_redirect_url
        # --- END URL CONSTRUCTION ---

        if 'save_region' in request.POST:
            form = RegionForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Region added successfully!") # Success message
                return redirect(final_redirect_url)
            else:
                messages.error(request, "Please correct the errors below.") # General error message for invalid form
                # If form is invalid, 'form' object already contains errors and will be passed to context
                # No redirect here, so form with errors remains on page

        elif 'update_region' in request.POST:
            region_id_to_update = request.POST.get('region_id')
            if region_id_to_update:
                try:
                    region_instance = get_object_or_404(Region, id=region_id_to_update)
                    edit_form = RegionForm(request.POST, instance=region_instance)
                    if edit_form.is_valid():
                        edit_form.save()
                        messages.success(request, "Region updated successfully!") # Success message
                        return redirect(final_redirect_url)
                    else:
                        messages.error(request, "Please correct the errors below for updating.") # General error message
                except (Region.DoesNotExist, ValueError):
                    messages.error(request, "Region not found for update or invalid ID provided.")
            else:
                messages.error(request, "No region ID provided for update.")

    return render(request, 'inventory/create_region.html', {
        'form': form,
        'edit_form': edit_form,
        'edit_id': edit_id,
        'regions': regions,
        'search_query': search_query,
    })

@login_required
def delete_region(request, pk):
    region = get_object_or_404(Region, pk=pk)
    region_name = region.name # Capture name for the message
    region.delete()
    messages.info(request, f"Region '{region_name}' deleted successfully.") # Info message for deletion

    # Preserve GET parameters on redirect after deletion
    base_redirect_url = reverse('create_region')
    current_get_params = request.GET.dict()
    if current_get_params:
        query_string = urlencode(current_get_params)
        final_redirect_url = f"{base_redirect_url}?{query_string}"
    else:
        final_redirect_url = base_redirect_url
    return redirect(final_redirect_url)

# ------------------ STORE --------------------
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from .models import Store, Region  # Ensure these are correctly imported
from .forms import StoreForm  # Ensure this is correctly imported

@login_required
def create_store(request):
    # --- Filter Logic Start ---
    region_filter = request.GET.get('region_filter')
    search_query = request.GET.get('search')

    stores = Store.objects.select_related('region').order_by('name')

    if region_filter:
        try:
            region_filter = int(region_filter)
            stores = stores.filter(region__id=region_filter)
        except (ValueError, TypeError):
            messages.error(request, "Invalid region filter provided.")
            region_filter = None  # Reset to avoid pre-selecting bad value

    if search_query:
        stores = stores.filter(name__icontains=search_query)
    # --- Filter Logic End ---

    regions = Region.objects.all().order_by('name')

    # Form handling for Create/Update
    form = StoreForm()
    edit_id = request.GET.get('edit_id')
    edit_form = None

    if edit_id:
        try:
            edit_store_instance = Store.objects.get(id=edit_id)
            edit_form = StoreForm(instance=edit_store_instance)
        except Store.DoesNotExist:
            messages.error(request, "Store not found for editing.")
            edit_id = None

    if request.method == 'POST':
        # Get all current GET parameters for redirection, excluding potential 'edit_id'
        # if we are doing a save/update that removes it.
        # We need to construct the GET parameters string correctly.
        query_params = request.GET.copy()  # Make a mutable copy of QueryDict

        # Remove 'edit_id' from query_params if it was present, as it won't be relevant
        # after a save/update action (unless it's an update failure)
        if 'edit_id' in query_params:
            del query_params['edit_id']

        # Ensure we construct the redirect URL with only the relevant filters
        redirect_query_string = query_params.urlencode()

        # This will be the base URL for redirection
        redirect_base_url = 'create_store'

        if 'save_store' in request.POST:
            form = StoreForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, f"Store '{form.instance.name}' added successfully!")
                # Redirect using the named URL and the constructed query string
                return redirect(
                    f"{redirect_base_url}?{redirect_query_string}" if redirect_query_string else redirect_base_url)
            else:
                messages.error(request, "Error adding store. Please correct the errors below.")
                if edit_id:  # Re-initialize edit_form if it was active
                    try:
                        edit_store_instance = Store.objects.get(id=edit_id)
                        edit_form = StoreForm(instance=edit_store_instance)
                    except Store.DoesNotExist:
                        edit_id = None

        elif 'update_store' in request.POST:
            store_id_to_update = request.POST.get('store_id')
            if store_id_to_update:
                try:
                    store = Store.objects.get(id=store_id_to_update)
                    edit_form = StoreForm(request.POST, instance=store)
                    if edit_form.is_valid():
                        edit_form.save()
                        messages.success(request, f"Store '{edit_form.instance.name}' updated successfully!")
                        # Redirect using the named URL and the constructed query string
                        return redirect(
                            f"{redirect_base_url}?{redirect_query_string}" if redirect_query_string else redirect_base_url)
                    else:
                        messages.error(request, "Error updating store. Please correct the errors below.")
                except Store.DoesNotExist:
                    messages.error(request, "Attempted to update a store that does not exist.")
                    # Redirect to clear invalid state, preserving filters
                    return redirect(
                        f"{redirect_base_url}?{redirect_query_string}" if redirect_query_string else redirect_base_url)

    return render(request, 'inventory/create_store.html', {
        'form': form,
        'edit_form': edit_form,
        'edit_id': edit_id,
        'stores': stores,
        'regions': regions,
        'selected_region': region_filter,
        'search_query': search_query,
    })

@login_required
def delete_store(request, pk):
    store = get_object_or_404(Store, pk=pk)
    store_name = store.name

    # Get all current GET parameters for redirection
    current_get_params = request.GET.urlencode()
    redirect_base_url = 'create_store'

    store.delete()
    messages.success(request, f"Store '{store_name}' deleted successfully!")
    # Redirect using the named URL and the constructed query string
    return redirect(f"{redirect_base_url}?{current_get_params}" if current_get_params else redirect_base_url)


# ------------------ FIELD OFFICER --------------------
@login_required
def create_field_officer(request):
    # --- Filter Logic Start ---
    store_filter = request.GET.get('store_filter')
    search_query = request.GET.get('search')

    field_officers = FieldOfficer.objects.select_related('store').order_by('name')

    if store_filter:
        try:
            store_filter = int(store_filter)
            field_officers = field_officers.filter(store__id=store_filter)
        except (ValueError, TypeError):
            messages.error(request, "Invalid store filter provided.")
            store_filter = None  # Reset to avoid pre-selecting a bad value

    if search_query:
        field_officers = field_officers.filter(name__icontains=search_query)
    # --- Filter Logic End ---

    stores_for_filter = Store.objects.all().order_by('name')

    # Form handling for Create/Update
    form = FieldOfficerForm()  # Form for new FO creation
    edit_id = request.GET.get('edit_id')  # ID for FO being edited
    edit_form = None  # Initialize edit_form

    if edit_id:
        try:
            edit_fo_instance = FieldOfficer.objects.get(id=edit_id)
            edit_form = FieldOfficerForm(instance=edit_fo_instance)
        except FieldOfficer.DoesNotExist:
            messages.error(request, "Field Officer not found for editing.")
            edit_id = None  # Reset to avoid errors if ID was bad

    if request.method == 'POST':
        # Get all current GET parameters for redirection, excluding 'edit_id'
        query_params = request.GET.copy()
        if 'edit_id' in query_params:
            del query_params['edit_id']
        redirect_query_string = query_params.urlencode()
        redirect_base_url = reverse('create_field_officer')

        if 'save_fo' in request.POST:
            form = FieldOfficerForm(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, f"Field Officer '{form.instance.name}' added successfully!")
                    return redirect(
                        f"{redirect_base_url}?{redirect_query_string}" if redirect_query_string else redirect_base_url)
                except IntegrityError: # Catch duplicate field_officer_code
                    messages.error(request, "Error adding Field Officer: The provided code is already in use. Please enter a unique code.")
                except Exception as e:  # Catch other potential errors
                    messages.error(request, f"Error adding Field Officer: {e}.")
            else:
                messages.error(request, "Error adding Field Officer. Please correct the errors below.")
                # If it was an 'add' attempt with a pre-existing edit_id,
                # re-initialize edit_form to keep editing context active
                if edit_id:
                    try:
                        edit_fo_instance = FieldOfficer.objects.get(id=edit_id)
                        edit_form = FieldOfficerForm(instance=edit_fo_instance)
                    except FieldOfficer.DoesNotExist:
                        edit_id = None

        elif 'update_fo' in request.POST:
            fo_id_to_update = request.POST.get('fo_id')
            if fo_id_to_update:
                try:
                    fo = FieldOfficer.objects.get(id=fo_id_to_update)
                    edit_form = FieldOfficerForm(request.POST, instance=fo)
                    if edit_form.is_valid():
                        try:
                            edit_form.save()
                            messages.success(request,
                                             f"Field Officer '{edit_form.instance.name}' updated successfully!")
                            return redirect(
                                f"{redirect_base_url}?{redirect_query_string}" if redirect_query_string else redirect_base_url)
                        except IntegrityError: # Catch duplicate field_officer_code
                            messages.error(request,
                                           "Error updating Field Officer: The provided code is already in use. Please enter a unique code.")
                        except Exception as e:  # Catch other potential errors
                            messages.error(request,
                                           f"Error updating Field Officer: {e}.")
                    else:
                        messages.error(request, "Error updating Field Officer. Please correct the errors below.")
                except FieldOfficer.DoesNotExist:
                    messages.error(request, "Attempted to update a Field Officer that does not exist.")
                    # Redirect to clear invalid state, preserving filters
                    return redirect(
                        f"{redirect_base_url}?{redirect_query_string}" if redirect_query_string else redirect_base_url)
            else:
                messages.error(request, "No Field Officer ID provided for update.")

    return render(request, 'inventory/create_field_officer.html', {
        'form': form,  # Form for creating new
        'edit_form': edit_form,  # Form for editing existing (if edit_id present)
        'edit_id': edit_id,  # ID of FO being edited, for template logic
        'field_officers': field_officers,  # Filtered list of FOs for display
        'stores_for_filter': stores_for_filter,  # All stores for the filter dropdown
        'selected_store': store_filter,  # Pass selected store ID back to pre-select filter
        'search_query': search_query,  # Pass search query back to pre-fill search box
    })

@login_required
def delete_field_officer(request, pk):
    fo = get_object_or_404(FieldOfficer, pk=pk)
    fo_name = fo.name  # Store name before deletion for message

    # Get current GET parameters from the URL (e.g. store_filter=2&search=jack)
    current_get_params = request.GET.urlencode()

    # ✅ Reverse the actual route name WITHOUT query params
    base_url = reverse('create_field_officer')

    try:
        fo.delete()
        messages.success(request, f"Field Officer '{fo_name}' deleted successfully!")
    except Exception as e:
        messages.error(request, f"Error deleting Field Officer '{fo_name}': {e}")

    # ✅ Append query parameters manually to preserve filters
    redirect_url = f"{base_url}?{current_get_params}" if current_get_params else base_url

    return redirect(redirect_url)
# ****************************************************************************************************************

@login_required
def edit_sample(request, pk):
    sample = get_object_or_404(ReceivingNormalSKR, pk=pk)
    sizing = get_object_or_404(Sizing, normal_skr_entry=sample)

    # You can create a Sizing form inline or via ModelForm
    SizingForm = modelform_factory(Sizing, fields=['size_0', 'size_1l', 'size_1s', 'size_1xs'])

    if request.method == 'POST':
        skr_form = ReceivingNormalSKRForm(request.POST, instance=sample)
        sizing_form = SizingForm(request.POST, instance=sizing)

        if skr_form.is_valid() and sizing_form.is_valid():
            skr_form.save()
            sizing_form.save()
            return redirect('receiving_detail', pk=sample.batch.pk)
    else:
        skr_form = ReceivingNormalSKRForm(instance=sample)
        sizing_form = SizingForm(instance=sizing)

    return render(request, 'inventory/edit_sample.html', {
        'skr_form': skr_form,
        'sizing_form': sizing_form,
        'sample': sample
    })

@login_required
def delete_sample(request, pk):
    sample = get_object_or_404(ReceivingNormalSKR, pk=pk)
    batch_id = sample.batch.pk
    sample.delete()
    return redirect('receiving_detail', pk=batch_id)


# from .models import Batch, ReceivingNormalSKR
@login_required
def download_normal_skr(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)
    samples = ReceivingNormalSKR.objects.filter(batch=batch).select_related(
        'field_officer__store__region'
    ).order_by(
        'field_officer__store__region__name',
        'field_officer__store__name',
        'field_officer__name',
        'sample_number'
    )

    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for sample in samples:
        region = sample.field_officer.store.region.name
        store = sample.field_officer.store.name
        officer = sample.field_officer.name
        grouped[region][store][officer].append(sample)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch {batch.batch_number}"

    row_offset = 2
    col_offset = 2
    data_start_col = col_offset + 1

    current_col = data_start_col
    sample_cells = []

    vibrant_green_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    shouting_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))

    base_font = Font(name='Baskerville Old Face', size=12)

    def set_cell_style(cell, value, fill=None, bold_text=False):
        cell.value = value
        cell.font = Font(name=base_font.name, size=base_font.size, bold=bold_text)
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill

    set_cell_style(ws.cell(row=row_offset, column=col_offset), "Field Officer", bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 1, column=col_offset), "Stores", bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 2, column=col_offset), "Region", bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 3, column=col_offset), "Samples", bold_text=True)

    region_col_ranges = defaultdict(lambda: {'start': None, 'end': None})
    store_col_ranges = defaultdict(lambda: {'start': None, 'end': None})
    officer_col_ranges = defaultdict(lambda: {'start': None, 'end': None})

    prev_region = prev_store = prev_officer = None

    for region, stores in grouped.items():
        if prev_region and region != prev_region and region_col_ranges[prev_region]['start'] is not None:
            region_col_ranges[prev_region]['end'] = current_col - 1

        if region_col_ranges[region]['start'] is None:
            region_col_ranges[region]['start'] = current_col

        for store, officers in stores.items():
            if prev_store and store != prev_store and store_col_ranges[prev_store]['start'] is not None:
                store_col_ranges[prev_store]['end'] = current_col - 1

            if store_col_ranges[store]['start'] is None:
                store_col_ranges[store]['start'] = current_col

            for officer, samples_list in officers.items():
                if prev_officer and officer != prev_officer and officer_col_ranges[prev_officer]['start'] is not None:
                    officer_col_ranges[prev_officer]['end'] = current_col - 1

                if officer_col_ranges[officer]['start'] is None:
                    officer_col_ranges[officer]['start'] = current_col

                set_cell_style(ws.cell(row=row_offset, column=current_col), officer, vibrant_green_fill, bold_text=True)
                set_cell_style(ws.cell(row=row_offset + 1, column=current_col), store, vibrant_green_fill,
                               bold_text=True)
                set_cell_style(ws.cell(row=row_offset + 2, column=current_col), region, vibrant_green_fill,
                               bold_text=True)

                for sample in samples_list:
                    set_cell_style(ws.cell(row=row_offset + 3, column=current_col), sample.sample_number,
                                   vibrant_green_fill, bold_text=True)
                    sample_cells.append(sample)
                    current_col += 1

                officer_col_ranges[officer]['end'] = current_col - 1
                prev_officer = officer

            store_col_ranges[store]['end'] = current_col - 1
            prev_store = store

        region_col_ranges[region]['end'] = current_col - 1
        prev_region = region

    for group, r in [(officer_col_ranges, row_offset), (store_col_ranges, row_offset + 1),
                     (region_col_ranges, row_offset + 2)]:
        for key, val in group.items():
            if val['start'] < val['end']:
                ws.merge_cells(start_row=r, start_column=val['start'], end_row=r, end_column=val['end'])

    metrics = [
        ("Good Q", lambda s: s.good_q, False),
        ("Insect Damaged", lambda s: s.insect_damaged, False),
        ("Mold", lambda s: s.mold, False),
        ("Immature", lambda s: s.immature, False),
        ("Weight", lambda s: s.weight, False),
        ("SKR", lambda s: s.skr, True),
        ("Moisture", lambda s: s.moisture, False),
        ("spillage", lambda s: s.spillage, False),
    ]

    metric_start_row = row_offset + 4
    sample_count = len(sample_cells)
    avg_col = data_start_col + sample_count
    weighted_col = avg_col + 1

    weight_row_index = None
    spillage_row_index = None

    for i, (label, extractor, is_formula) in enumerate(metrics):
        row = metric_start_row + i
        set_cell_style(ws.cell(row=row, column=col_offset), label,
                       shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold", "Immature"] else None,
                       bold_text=True)

        for j, sample in enumerate(sample_cells):
            cell = ws.cell(row=row, column=data_start_col + j)
            if is_formula and label == "SKR":
                good_q_row = metric_start_row + 0
                good_q_cell = f"{get_column_letter(data_start_col + j)}{good_q_row}"
                formula = f"=IFERROR({good_q_cell}/5, 0)"
                set_cell_style(cell, formula, bold_text=False)
            else:
                val = extractor(sample)
                set_cell_style(cell, float(val) if val is not None else 0,
                               shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold",
                                                                 "Immature"] else None)
            cell.number_format = '0.00'

        if label.lower() == "weight":
            weight_row_index = row
        if label.lower() == "spillage":
            spillage_row_index = row

        if sample_count > 0:
            start_letter = get_column_letter(data_start_col)
            end_letter = get_column_letter(data_start_col + sample_count - 1)
            avg_cell = ws.cell(row=row, column=avg_col)
            formula = f"=AVERAGE({start_letter}{row}:{end_letter}{row})"
            set_cell_style(avg_cell, formula,
                           shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold",
                                                             "Immature"] else vibrant_green_fill if label == "SKR" else None)
            avg_cell.number_format = '0.00'

    for i, (label, _, _) in enumerate(metrics):
        row = metric_start_row + i
        weighted_cell = ws.cell(row=row, column=weighted_col)
        if label not in ["Weight", "Good Q", "Insect Damaged", "Mold", "Immature", "SKR", "Moisture"]:
            set_cell_style(weighted_cell, "N/A", bold_text=True)
        else:
            terms = []
            for j in range(sample_count):
                metric_col = get_column_letter(data_start_col + j)
                weight_ref = f"${metric_col}${weight_row_index}"
                metric_ref = f"{metric_col}{row}"
                terms.append(f"({metric_ref}*{weight_ref})")
            numerator = "+".join(terms)
            weight_start = get_column_letter(data_start_col)
            weight_end = get_column_letter(data_start_col + sample_count - 1)
            total_weight = f"SUM(${weight_start}${weight_row_index}:${weight_end}${weight_row_index})"
            spillage_value = "0"
            if spillage_row_index:
                spillage_avg_ref = f"{get_column_letter(avg_col)}{spillage_row_index}"
                spillage_value = f"IF(ISBLANK({spillage_avg_ref}), 0, {spillage_avg_ref})"
            denominator = f"({total_weight} - {spillage_value})"
            formula = f"=IF({denominator}=0, \"N/A\", ({numerator})/({denominator}))"
            set_cell_style(weighted_cell, formula,
                           shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold",
                                                             "Immature"] else vibrant_green_fill if label == "SKR" else None)
            weighted_cell.number_format = '0.00'

    set_cell_style(ws.cell(row=row_offset + 3, column=avg_col), "Average", vibrant_green_fill, bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 3, column=weighted_col), "Weighted", vibrant_green_fill, bold_text=True)

    for r in range(row_offset, metric_start_row + len(metrics)):
        cell = ws.cell(row=r, column=col_offset)
        cell.border = thick_border
        cell.font = Font(name=base_font.name, size=base_font.size, bold=True)
        if r in [row_offset, row_offset + 1, row_offset + 2, row_offset + 3]:
            cell.fill = vibrant_green_fill
        elif r in [metric_start_row + i for i, (label, _, _) in enumerate(metrics) if
                   label in ["Good Q", "Insect Damaged", "Mold", "Immature"]]:
            cell.fill = shouting_yellow_fill

    for col in range(col_offset, weighted_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions[get_column_letter(1)].width = 3

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"NormalSKR_Merged_Batch_{batch.batch_number}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ************MositureSKR**************
@login_required
def add_moisture_skr(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    regions = Region.objects.all()

    # Initialize variables to hold selected values for re-rendering form on error
    selected_region_id = None
    selected_store_id = None
    fo_id = None
    stores_in_region = []
    field_officers_in_store = []
    form_data = request.POST # To re-populate form fields on error

    if request.method == 'POST':
        # Retrieve field officer ID from the hidden input 'fo'
        fo_id = request.POST.get('fo')

        if not fo_id:
            errors = ["Please select a Field Officer."]
            # Attempt to re-populate regions, stores, and FOs for error display
            if request.POST.get('region'):
                selected_region_id = request.POST.get('region')
                stores_in_region = Store.objects.filter(region_id=selected_region_id)
            if request.POST.get('store'):
                selected_store_id = request.POST.get('store')
                field_officers_in_store = FieldOfficer.objects.filter(store_id=selected_store_id)

            return render(request, 'inventory/add_moisture_skr.html', {
                'batch': batch,
                'regions': regions,
                'errors': errors,
                'selected_region_id': selected_region_id,
                'selected_store_id': selected_store_id,
                'fo_id': fo_id,
                'stores_in_region': stores_in_region,
                'field_officers_in_store': field_officers_in_store,
                'form_data': form_data, # Pass all POST data to repopulate sample fields
            })

        field_officer = get_object_or_404(FieldOfficer, id=fo_id)

        good_q_list = request.POST.getlist('good_q[]')
        insect_damaged_list = request.POST.getlist('insect_damaged[]')
        mold_list = request.POST.getlist('mold[]')
        immature_list = request.POST.getlist('immature[]')
        weight_list = request.POST.getlist('weight[]')
        moisture_list = request.POST.getlist('moisture[]')
        spillage_list = request.POST.getlist('spillage[]')

        errors = []
        samples_to_save = []

        # Determine the starting sample number for this submission
        last_sample = ReceivingSKRWithMoisture.objects.filter(
            batch=batch, field_officer=field_officer
        ).aggregate(Max('sample_number'))['sample_number__max'] or 0

        for i in range(len(good_q_list)):
            try:
                # Use .strip() to handle potential whitespace from user input
                good_q = Decimal(good_q_list[i].strip() if good_q_list[i].strip() else '0')
                insect_damaged = Decimal(insect_damaged_list[i].strip() if insect_damaged_list[i].strip() else '0')
                mold = Decimal(mold_list[i].strip() if mold_list[i].strip() else '0')
                immature = Decimal(immature_list[i].strip() if immature_list[i].strip() else '0')

                total = good_q + insect_damaged + mold + immature

                if abs(total - 100) > Decimal('0.1'): # Allowing a small tolerance for floating point
                    errors.append(f"Sample #{i+1}: Composition adds to {total}%. Must equal 100%.")
                else:
                    weight = Decimal(weight_list[i].strip() if weight_list[i].strip() else '0')
                    moisture = Decimal(moisture_list[i].strip() if moisture_list[i].strip() else '0')
                    spillage = Decimal(spillage_list[i].strip() if spillage_list[i].strip() else '0')
                    skr = round(good_q / Decimal(5), 2)

                    samples_to_save.append({
                        'sample_number': last_sample + i + 1, # Increment sample number correctly
                        'good_q': good_q,
                        'insect_damaged': insect_damaged,
                        'mold': mold,
                        'immature': immature,
                        'weight': weight,
                        'moisture': moisture,
                        'spillage': spillage,
                        'skr': skr,
                    })

            except (ValueError, InvalidOperation):
                errors.append(f"Sample #{i+1}: Invalid number format in one or more fields.")
            except IndexError:
                errors.append(f"Sample #{i+1}: Missing data for some fields.")


        if errors:
            # Re-populate selected region, store, and FO for display
            selected_region_id = request.POST.get('region')
            selected_store_id = request.POST.get('store')
            # fo_id is already set from the POST request at the beginning

            if selected_region_id:
                stores_in_region = Store.objects.filter(region_id=selected_region_id)
            if selected_store_id:
                field_officers_in_store = FieldOfficer.objects.filter(store_id=selected_store_id)

            return render(request, 'inventory/add_moisture_skr.html', {
                'batch': batch,
                'regions': regions,
                'errors': errors,
                'selected_region_id': selected_region_id,
                'selected_store_id': selected_store_id,
                'fo_id': fo_id,
                'stores_in_region': stores_in_region,
                'field_officers_in_store': field_officers_in_store,
                'form_data': form_data, # Pass all POST data to repopulate sample fields
            })

        # Save only if all samples passed validation
        with transaction.atomic():
            for sample_data in samples_to_save:
                ReceivingSKRWithMoisture.objects.create(
                    batch=batch,
                    field_officer=field_officer,
                    sample_number=sample_data['sample_number'],
                    good_q=sample_data['good_q'],
                    insect_damaged=sample_data['insect_damaged'],
                    mold=sample_data['mold'],
                    immature=sample_data['immature'],
                    weight=sample_data['weight'],
                    moisture=sample_data['moisture'],
                    spillage=sample_data['spillage'],
                    skr=sample_data['skr'],
                )
        # Assuming 'receiving_detail' is the correct redirect URL after successful submission
        return redirect('receiving_detail', pk=batch.pk)

    # Initial GET request
    return render(request, 'inventory/add_moisture_skr.html', {
        'batch': batch,
        'regions': regions,
        'selected_region_id': selected_region_id,
        'selected_store_id': selected_store_id,
        'fo_id': fo_id,
        'stores_in_region': stores_in_region,
        'field_officers_in_store': field_officers_in_store,
    })


@login_required
def edit_moisture_skr(request, pk):
    sample = get_object_or_404(ReceivingSKRWithMoisture, pk=pk)
    if request.method == 'POST':
        form = ReceivingSKRWithMoistureForm(request.POST, instance=sample)
        if form.is_valid():
            form.save()
            return redirect('receiving_detail', pk=sample.batch.pk)
    else:
        # CORRECTED LINE: Use your form class here, not the model class
        form = ReceivingSKRWithMoistureForm(instance=sample)
    return render(request, 'inventory/edit_moisture_sample.html', {'form': form, 'sample': sample})

@login_required
def delete_moisture_skr(request, pk):
    sample = get_object_or_404(ReceivingSKRWithMoisture, pk=pk)
    batch_id = sample.batch.pk
    sample.delete()
    return redirect('receiving_detail', pk=batch_id)


from .models import ReceivingSKRWithMoisture

@login_required
def download_skr_moisture(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)
    samples = ReceivingSKRWithMoisture.objects.filter(batch=batch).select_related(
        'field_officer__store__region'
    ).order_by(
        'field_officer__store__region__name',
        'field_officer__store__name',
        'field_officer__name',
        'sample_number'
    )

    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for sample in samples:
        region = sample.field_officer.store.region.name
        store = sample.field_officer.store.name
        officer = sample.field_officer.name
        grouped[region][store][officer].append(sample)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch {batch.batch_number}"

    row_offset = 2
    col_offset = 2
    data_start_col = col_offset + 1

    current_col = data_start_col
    sample_cells = []

    vibrant_green_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    shouting_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))

    base_font = Font(name='Baskerville Old Face', size=12)

    def set_cell_style(cell, value, fill=None, bold_text=False):
        cell.value = value
        cell.font = Font(name=base_font.name, size=base_font.size, bold=bold_text)
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill

    set_cell_style(ws.cell(row=row_offset, column=col_offset), "Field Officer", bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 1, column=col_offset), "Stores", bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 2, column=col_offset), "Region", bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 3, column=col_offset), "Samples", bold_text=True)

    region_col_ranges = defaultdict(lambda: {'start': None, 'end': None})
    store_col_ranges = defaultdict(lambda: {'start': None, 'end': None})
    officer_col_ranges = defaultdict(lambda: {'start': None, 'end': None})

    prev_region = prev_store = prev_officer = None

    for region, stores in grouped.items():
        if prev_region and region != prev_region and region_col_ranges[prev_region]['start'] is not None:
            region_col_ranges[prev_region]['end'] = current_col - 1

        if region_col_ranges[region]['start'] is None:
            region_col_ranges[region]['start'] = current_col

        for store, officers in stores.items():
            if prev_store and store != prev_store and store_col_ranges[prev_store]['start'] is not None:
                store_col_ranges[prev_store]['end'] = current_col - 1

            if store_col_ranges[store]['start'] is None:
                store_col_ranges[store]['start'] = current_col

            for officer, samples_list in officers.items():
                if prev_officer and officer != prev_officer and officer_col_ranges[prev_officer]['start'] is not None:
                    officer_col_ranges[prev_officer]['end'] = current_col - 1

                if officer_col_ranges[officer]['start'] is None:
                    officer_col_ranges[officer]['start'] = current_col

                set_cell_style(ws.cell(row=row_offset, column=current_col), officer, vibrant_green_fill, bold_text=True)
                set_cell_style(ws.cell(row=row_offset + 1, column=current_col), store, vibrant_green_fill,
                               bold_text=True)
                set_cell_style(ws.cell(row=row_offset + 2, column=current_col), region, vibrant_green_fill,
                               bold_text=True)

                for sample in samples_list:
                    set_cell_style(ws.cell(row=row_offset + 3, column=current_col), sample.sample_number,
                                   vibrant_green_fill, bold_text=True)
                    sample_cells.append(sample)
                    current_col += 1

                officer_col_ranges[officer]['end'] = current_col - 1
                prev_officer = officer

            store_col_ranges[store]['end'] = current_col - 1
            prev_store = store

        region_col_ranges[region]['end'] = current_col - 1
        prev_region = region

    for group, r in [(officer_col_ranges, row_offset), (store_col_ranges, row_offset + 1),
                     (region_col_ranges, row_offset + 2)]:
        for key, val in group.items():
            if val['start'] < val['end']:
                ws.merge_cells(start_row=r, start_column=val['start'], end_row=r, end_column=val['end'])

    metrics = [
        ("Good Q", lambda s: s.good_q, False),
        ("Insect Damaged", lambda s: s.insect_damaged, False),
        ("Mold", lambda s: s.mold, False),
        ("Immature", lambda s: s.immature, False),
        ("Weight", lambda s: s.weight, False),
        ("Moisture", lambda s: s.moisture, False),
        ("SKR", lambda s: s.skr, True),
        ("spillage", lambda s: s.spillage, False),
    ]

    metric_start_row = row_offset + 4
    sample_count = len(sample_cells)
    avg_col = data_start_col + sample_count
    weighted_col = avg_col + 1

    weight_row_index = None
    spillage_row_index = None
    moisture_row_index = None  # New: To store the row index for Moisture

    # First pass to find row indices for 'Weight' and 'Moisture'
    for i, (label, _, _) in enumerate(metrics):
        row = metric_start_row + i
        if label.lower() == "weight":
            weight_row_index = row
        if label.lower() == "spillage":
            spillage_row_index = row
        if label.lower() == "moisture":
            moisture_row_index = row

    for i, (label, extractor, is_formula) in enumerate(metrics):
        row = metric_start_row + i
        set_cell_style(ws.cell(row=row, column=col_offset), label,
                       shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold", "Immature"] else None,
                       bold_text=True)

        for j, sample in enumerate(sample_cells):
            cell = ws.cell(row=row, column=data_start_col + j)
            if is_formula and label == "SKR":
                good_q_cell = f"{get_column_letter(data_start_col + j)}{metric_start_row + 0}"  # Good Q is the first metric
                moisture_avg_cell = f"{get_column_letter(avg_col)}{moisture_row_index}"  # Average of Moisture
                # Updated SKR formula
                formula = f"=IFERROR({good_q_cell}/(100/(100-55-1.5-{moisture_avg_cell})), 0)"
                set_cell_style(cell, formula, bold_text=False)
            else:
                val = extractor(sample)
                set_cell_style(cell, float(val) if val is not None else 0,
                               shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold",
                                                                 "Immature"] else None)
            cell.number_format = '0.00'

        if sample_count > 0:
            start_letter = get_column_letter(data_start_col)
            end_letter = get_column_letter(data_start_col + sample_count - 1)
            avg_cell = ws.cell(row=row, column=avg_col)
            formula = f"=AVERAGE({start_letter}{row}:{end_letter}{row})"
            set_cell_style(avg_cell, formula,
                           shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold",
                                                             "Immature"] else vibrant_green_fill if label == "SKR" else None)
            avg_cell.number_format = '0.00'

    for i, (label, _, _) in enumerate(metrics):
        row = metric_start_row + i
        weighted_cell = ws.cell(row=row, column=weighted_col)
        if label not in ["Weight", "Good Q", "Insect Damaged", "Mold", "Immature", "SKR", "Moisture"]:
            set_cell_style(weighted_cell, "N/A", bold_text=True)
        else:
            terms = []
            for j in range(sample_count):
                metric_col = get_column_letter(data_start_col + j)
                weight_ref = f"${metric_col}${weight_row_index}"
                metric_ref = f"{metric_col}{row}"
                terms.append(f"({metric_ref}*{weight_ref})")
            numerator = "+".join(terms)
            weight_start = get_column_letter(data_start_col)
            weight_end = get_column_letter(data_start_col + sample_count - 1)
            total_weight = f"SUM(${weight_start}${weight_row_index}:${weight_end}${weight_row_index})"
            spillage_value = "0"
            if spillage_row_index:
                spillage_avg_ref = f"{get_column_letter(avg_col)}{spillage_row_index}"
                spillage_value = f"IF(ISBLANK({spillage_avg_ref}), 0, {spillage_avg_ref})"
            denominator = f"({total_weight} - {spillage_value})"
            formula = f"=IF({denominator}=0, \"N/A\", ({numerator})/({denominator}))"
            set_cell_style(weighted_cell, formula,
                           shouting_yellow_fill if label in ["Good Q", "Insect Damaged", "Mold",
                                                             "Immature"] else vibrant_green_fill if label == "SKR" else None)
            weighted_cell.number_format = '0.00'

    set_cell_style(ws.cell(row=row_offset + 3, column=avg_col), "Average", vibrant_green_fill, bold_text=True)
    set_cell_style(ws.cell(row=row_offset + 3, column=weighted_col), "Weighted", vibrant_green_fill, bold_text=True)

    for r in range(row_offset, metric_start_row + len(metrics)):
        cell = ws.cell(row=r, column=col_offset)
        cell.border = thick_border
        cell.font = Font(name=base_font.name, size=base_font.size, bold=True)
        if r in [row_offset, row_offset + 1, row_offset + 2, row_offset + 3]:
            cell.fill = vibrant_green_fill
        elif r in [metric_start_row + i for i, (label, _, _) in enumerate(metrics) if
                   label in ["Good Q", "Insect Damaged", "Mold", "Immature"]]:
            cell.fill = shouting_yellow_fill

    for col in range(col_offset, weighted_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions[get_column_letter(1)].width = 3

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"SKR_Moisture_Batch_{batch.batch_number}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ************************************************************************************************************
import io
from decimal import InvalidOperation


# Assuming these models are defined and imported correctly from your app's models.py
# from .models import Batch, Region, FieldOfficer, ReceivingNormalSKR, Sizing

@login_required
def download_fo_average(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)
    samples = ReceivingNormalSKR.objects.filter(batch=batch).select_related('field_officer__store__region')

    print(f"DEBUG: Starting download_fo_average for batch ID: {batch_id}")
    print(f"DEBUG: Found {samples.count()} samples for batch ID {batch_id}")
    if not samples.exists():
        print("DEBUG: No samples found for this batch. Excel will likely be empty.")

    fo_raw_sample_values = defaultdict(lambda: defaultdict(list))
    fo_region_map = {}

    for sample in samples:
        fo = sample.field_officer
        if not fo:
            print(f"DEBUG: Sample ID {sample.id} has no associated Field Officer. Skipping.")
            continue

        fo_name = fo.name
        region_name = fo.store.region.name if fo.store and fo.store.region else "Unknown Region"
        fo_region_map[fo_name] = region_name

        metrics_to_process = [
            ('good_q', 'Good Q'),
            ('insect_damaged', 'Insect Damaged'),
            ('mold', 'Mold'),
            ('immature', 'Immature'),
            ('weight', 'Weight'),
            ('moisture', 'Moisture'),
        ]

        sample_has_data = False
        for attr_key, display_key in metrics_to_process:
            value = getattr(sample, attr_key, None)
            if value is not None:
                try:
                    decimal_value = Decimal(str(value))
                    fo_raw_sample_values[fo_name][display_key].append(decimal_value)
                    sample_has_data = True
                except InvalidOperation:
                    print(
                        f"DEBUG: Warning: Could not convert '{value}' from sample {sample.id}, attribute {attr_key} to Decimal.")

        if sample.good_q is not None:
            try:
                skr_value = Decimal(str(sample.good_q)) / Decimal(5)
                fo_raw_sample_values[fo_name]['SKR'].append(skr_value)
                sample_has_data = True
            except (InvalidOperation, ZeroDivisionError) as e:
                print(f"DEBUG: Warning: Could not calculate SKR for sample {sample.id} (good_q={sample.good_q}): {e}")

        if not sample_has_data:
            print(f"DEBUG: No numeric data (or SKR) collected for sample ID {sample.id} (FO: {fo_name}).")

    print(f"DEBUG: Raw sample values collected per FO: {fo_raw_sample_values}")

    fo_calculated_averages = defaultdict(dict)
    for fo_name, metrics_data in fo_raw_sample_values.items():
        for metric_display_name, values in metrics_data.items():
            if values:
                try:
                    calculated_mean = mean(values)
                    fo_calculated_averages[fo_name][metric_display_name] = round(Decimal(str(calculated_mean)), 2)
                except Exception as e:
                    print(f"DEBUG: Error calculating mean for FO {fo_name} - Metric {metric_display_name}: {e}")
                    fo_calculated_averages[fo_name][metric_display_name] = None
            else:
                fo_calculated_averages[fo_name][metric_display_name] = None

    print(f"DEBUG: Calculated FO Averages: {fo_calculated_averages}")

    # --- DYNAMIC FO AND REGION ORDERING ---
    # Get unique FOs and sort them alphabetically
    present_individual_fo_order = sorted(list(fo_raw_sample_values.keys()))

    # Get unique regions from the FOs present and sort them alphabetically
    present_aggregate_region_order = sorted(list(set(fo_region_map.values())))

    print(f"DEBUG: FOs to display: {present_individual_fo_order}")
    print(f"DEBUG: Regions to display: {present_aggregate_region_order}")

    wb = Workbook()
    ws = wb.active
    ws.title = "FO's Average"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    yellow_data_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    blue_aggregate_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    pink_aggregate_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    def set_cell(ws, row, col, value, bold_text=False, fill=None, number_format='0.00'):
        cell = ws.cell(row=row, column=col)
        # --- TEMPORARY DEBUG PRINT inside set_cell ---
        # print(f"DEBUG: set_cell R{row}C{col} received value: '{value}' (Type: {type(value)})")
        # --- END TEMPORARY DEBUG PRINT ---

        if value is None or value == "":
            cell.value = None
        else:
            cell.value = value

        cell.font = Font(bold=bold_text)
        cell.alignment = align_center
        cell.border = thin_border
        if fill:
            cell.fill = fill

        if isinstance(value, (float, int, Decimal)) or (isinstance(value, str) and str(value).startswith('=')):
            cell.number_format = number_format
        else:
            cell.number_format = 'General'

    total_columns_in_table = 1 + len(present_individual_fo_order) + len(present_aggregate_region_order)
    set_cell(ws, 1, 1, "Fo's Average", bold_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns_in_table)

    set_cell(ws, 2, 1, "FO", bold_text=True, fill=green_header_fill)

    fo_col_map = {}
    current_col_for_fo_headers = 2

    for fo_name in present_individual_fo_order:
        set_cell(ws, 2, current_col_for_fo_headers, fo_name, bold_text=True, fill=green_header_fill)
        fo_col_map[fo_name] = current_col_for_fo_headers
        current_col_for_fo_headers += 1

    if present_aggregate_region_order:
        aggregate_header_start_col = current_col_for_fo_headers
        aggregate_header_end_col = current_col_for_fo_headers + len(present_aggregate_region_order) - 1
        set_cell(ws, 2, aggregate_header_start_col, "Aggregate by region", bold_text=True, fill=green_header_fill)
        ws.merge_cells(start_row=2, start_column=aggregate_header_start_col, end_row=2,
                       end_column=aggregate_header_end_col)
        for col_in_range in range(aggregate_header_start_col, aggregate_header_end_col + 1):
            ws.cell(row=2, column=col_in_range).border = thin_border

    set_cell(ws, 3, 1, "Region", bold_text=True, fill=green_header_fill)

    current_fo_col_for_merge = 2
    fo_regions_in_display_order = [fo_region_map.get(fo, "Unknown") for fo in present_individual_fo_order]

    region_merge_ranges = []
    if fo_regions_in_display_order:
        current_merge_region_name = fo_regions_in_display_order[0]
        current_merge_start_col = 2

        for i, region_name_for_fo in enumerate(fo_regions_in_display_order):
            if region_name_for_fo != current_merge_region_name:
                region_merge_ranges.append({
                    'name': current_merge_region_name,
                    'start_col': current_merge_start_col,
                    'end_col': current_fo_col_for_merge - 1
                })
                current_merge_region_name = region_name_for_fo
                current_merge_start_col = current_fo_col_for_merge

            current_fo_col_for_merge += 1

        region_merge_ranges.append({
            'name': current_merge_region_name,
            'start_col': current_merge_start_col,
            'end_col': current_fo_col_for_merge - 1
        })

    for span_info in region_merge_ranges:
        start_col = span_info['start_col']
        end_col = span_info['end_col']

        set_cell(ws, 3, start_col, span_info['name'], bold_text=False, fill=green_header_fill)
        if start_col != end_col:
            ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)

        for col_in_range in range(start_col, end_col + 1):
            ws.cell(row=3, column=col_in_range).border = thin_border

    current_col_for_aggregate_region_label = 1 + len(present_individual_fo_order) + 1
    for i, region_name_for_aggregate_col in enumerate(present_aggregate_region_order):
        fill_color = blue_aggregate_fill if i % 2 == 0 else pink_aggregate_fill
        set_cell(ws, 3, current_col_for_aggregate_region_label, region_name_for_aggregate_col, bold_text=True,
                 fill=fill_color)
        current_col_for_aggregate_region_label += 1

    metrics_display_order = ["Good Q", "Insect Damaged", "Mold", "Immature", "Weight", "SKR", "Moisture"]

    for row_idx, metric_display_name in enumerate(metrics_display_order, start=4):
        set_cell(ws, row_idx, 1, metric_display_name, bold_text=True)
        current_col_for_data = 2

        for fo_name in present_individual_fo_order:
            avg_value = fo_calculated_averages[fo_name].get(metric_display_name)
            set_cell(ws, row_idx, current_col_for_data, avg_value, bold_text=True, fill=yellow_data_fill,
                     number_format='0.00')
            current_col_for_data += 1

        for i, target_region in enumerate(present_aggregate_region_order):
            fill_color = blue_aggregate_fill if i % 2 == 0 else pink_aggregate_fill

            fo_values_in_region = []
            for fo_name_in_order in present_individual_fo_order:
                if fo_region_map.get(fo_name_in_order) == target_region:
                    val = fo_calculated_averages[fo_name_in_order].get(metric_display_name)
                    if isinstance(val, (Decimal, float, int)):
                        fo_values_in_region.append(val)

            calculated_aggregate_value = None

            if fo_values_in_region:
                if metric_display_name == "Weight":
                    calculated_aggregate_value = sum(fo_values_in_region)
                else:
                    calculated_aggregate_value = mean(fo_values_in_region)

            if calculated_aggregate_value is not None:
                calculated_aggregate_value = round(Decimal(str(calculated_aggregate_value)), 2)

            set_cell(ws, row_idx, current_col_for_data, calculated_aggregate_value, bold_text=True, fill=fill_color,
                     number_format='0.00')
            current_col_for_data += 1

    for col in range(1, total_columns_in_table + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"FOs_Average_Batch_{batch.batch_number}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response


# SIZING
@login_required
def download_sizing_data_excel(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)

    # Fetch sizing data, linking through ReceivingNormalSKR to get FieldOfficer and Batch
    sizing_data = Sizing.objects.filter(
        normal_skr_entry__batch=batch
    ).select_related(
        'normal_skr_entry__field_officer'
    ).order_by(
        'normal_skr_entry__field_officer__name',
        'normal_skr_entry__sample_number'  # Order by sample number for consistent display
    )

    # Process data to group by Field Officer and calculate averages
    fo_sizing_data = defaultdict(lambda: {
        'samples': [],
        'size_0_total': Decimal(0), 'size_1l_total': Decimal(0),
        'size_1s_total': Decimal(0), 'size_1xs_total': Decimal(0),
        'sample_count': 0
    })

    all_sizes_total = defaultdict(Decimal)
    grand_sample_count = 0

    for entry in sizing_data:
        fo_name = entry.normal_skr_entry.field_officer.name
        sample_num = entry.normal_skr_entry.sample_number

        sample_sizing = {
            'sample_num': sample_num,
            'size_0': entry.size_0,
            'size_1l': entry.size_1l,
            'size_1s': entry.size_1s,
            'size_1xs': entry.size_1xs,
        }
        fo_sizing_data[fo_name]['samples'].append(sample_sizing)

        fo_sizing_data[fo_name]['size_0_total'] += entry.size_0
        fo_sizing_data[fo_name]['size_1l_total'] += entry.size_1l
        fo_sizing_data[fo_name]['size_1s_total'] += entry.size_1s
        fo_sizing_data[fo_name]['size_1xs_total'] += entry.size_1xs
        fo_sizing_data[fo_name]['sample_count'] += 1

        all_sizes_total['size_0'] += entry.size_0
        all_sizes_total['size_1l'] += entry.size_1l
        all_sizes_total['size_1s'] += entry.size_1s
        all_sizes_total['size_1xs'] += entry.size_1xs
        grand_sample_count += 1

    # Calculate FO averages
    for fo_name, data in fo_sizing_data.items():
        if data['sample_count'] > 0:
            data['averages'] = {
                'size_0': (data['size_0_total'] / data['sample_count']).quantize(Decimal('0.01')),
                'size_1l': (data['size_1l_total'] / data['sample_count']).quantize(Decimal('0.01')),
                'size_1s': (data['size_1s_total'] / data['sample_count']).quantize(Decimal('0.01')),
                'size_1xs': (data['size_1xs_total'] / data['sample_count']).quantize(Decimal('0.01')),
            }
        else:
            data['averages'] = {
                'size_0': Decimal('0.00'), 'size_1l': Decimal('0.00'),
                'size_1s': Decimal('0.00'), 'size_1xs': Decimal('0.00')
            }

    # Calculate Grand Averages
    grand_sizing_averages = {}
    if grand_sample_count > 0:
        grand_sizing_averages = {
            'size_0': (all_sizes_total['size_0'] / grand_sample_count).quantize(Decimal('0.01')),
            'size_1l': (all_sizes_total['size_1l'] / grand_sample_count).quantize(Decimal('0.01')),
            'size_1s': (all_sizes_total['size_1s'] / grand_sample_count).quantize(Decimal('0.01')),
            'size_1xs': (all_sizes_total['size_1xs'] / grand_sample_count).quantize(Decimal('0.01')),
        }
    else:
        grand_sizing_averages = {
            'size_0': Decimal('0.00'), 'size_1l': Decimal('0.00'),
            'size_1s': Decimal('0.00'), 'size_1xs': Decimal('0.00')
        }

    field_officer_names = sorted(fo_sizing_data.keys())
    max_samples_per_fo = 0
    if fo_sizing_data:
        max_samples_per_fo = max(len(data['samples']) for data in fo_sizing_data.values())

    # --- Excel Generation ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sizing_Batch_{batch.batch_number}"

    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    left_aligned_text = Alignment(horizontal="left", vertical="center")

    # Header Row 1 (FO Names)
    current_col = 2  # Start from column B
    ws.cell(row=1, column=1, value="FO").fill = grey_fill
    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=1, column=1).alignment = center_aligned_text
    ws.cell(row=1, column=1).border = thin_border

    for fo_name in field_officer_names:
        start_col = current_col
        end_col = start_col + max_samples_per_fo - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=fo_name)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_aligned_text
        for col_idx in range(start_col, end_col + 1):
            ws.cell(row=1, column=col_idx).border = thin_border  # Apply border to all merged cells
        current_col = end_col + 1

    # Grand Average Header
    ws.cell(row=1, column=current_col, value="Average").fill = yellow_fill
    ws.cell(row=1, column=current_col).font = bold_font
    ws.cell(row=1, column=current_col).alignment = center_aligned_text
    ws.cell(row=1, column=current_col).border = thin_border
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)

    # Header Row 2 (Sample Numbers)
    current_col = 2
    ws.cell(row=2, column=1, value="").fill = grey_fill  # Empty cell under "FO"
    ws.cell(row=2, column=1).border = thin_border
    for fo_name in field_officer_names:
        fo_data = fo_sizing_data[fo_name]
        for i, sample in enumerate(fo_data['samples']):
            cell = ws.cell(row=2, column=current_col, value=sample['sample_num'])
            cell.fill = grey_fill
            cell.font = bold_font
            cell.alignment = center_aligned_text
            cell.border = thin_border
            current_col += 1
        # Fill blank cells for FOs with fewer samples
        for _ in range(max_samples_per_fo - len(fo_data['samples'])):
            cell = ws.cell(row=2, column=current_col, value="")
            cell.fill = grey_fill
            cell.alignment = center_aligned_text
            cell.border = thin_border
            current_col += 1

    # Data Rows (Sizing values)
    sizing_types = ['size_0', 'size_1l', 'size_1s', 'size_1xs']
    sizing_labels = ['0', '1L', '1S', '1XS']

    for row_idx, size_type in enumerate(sizing_types, start=3):
        ws.cell(row=row_idx, column=1, value=sizing_labels[row_idx - 3]).fill = grey_fill
        ws.cell(row=row_idx, column=1).font = bold_font
        ws.cell(row=row_idx, column=1).alignment = left_aligned_text
        ws.cell(row=row_idx, column=1).border = thin_border

        current_col = 2
        for fo_name in field_officer_names:
            fo_data = fo_sizing_data[fo_name]
            for sample in fo_data['samples']:
                # Accessing dictionary values using bracket notation
                cell = ws.cell(row=row_idx, column=current_col, value=sample.get(size_type, ''))
                cell.alignment = center_aligned_text
                cell.border = thin_border
                current_col += 1
            # Fill blank cells for FOs with fewer samples
            for _ in range(max_samples_per_fo - len(fo_data['samples'])):
                cell = ws.cell(row=row_idx, column=current_col, value="")
                cell.alignment = center_aligned_text
                cell.border = thin_border
                current_col += 1

        # Grand Average Column
        cell = ws.cell(row=row_idx, column=current_col, value=grand_sizing_averages.get(size_type, Decimal('0.00')))
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_aligned_text
        cell.border = thin_border

    # Adjust column widths
    ws.column_dimensions[get_column_letter(1)].width = 10  # For the "FO" / Sizing Labels column
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    # Prepare HTTP response for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Sizing_Report{batch.batch_number}.xlsx'
    wb.save(response)
    return response


# ****************************** Final product****************
from .helpers import render_normal_skr_section, render_skr_with_moisture_section  # Step 3
@login_required
def download_combined_skr_excel(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch {batch.batch_number}"

    # Section header: Normal SKR
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=10)
    ws.cell(row=1, column=2).value = "NORMAL SKR SECTION"

    current_row = 2
    current_row = render_normal_skr_section(ws, batch, current_row)

    # Leave 4 blank rows
    current_row += 4

    # Section header: SKR With Moisture
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
    ws.cell(row=current_row, column=2).value = "SKR WITH MOISTURE SECTION"

    current_row += 1
    current_row = render_skr_with_moisture_section(ws, batch, current_row)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Merged_Batch_{batch.batch_number}.xlsx"'
    wb.save(response)
    return response
# *************************************************************************************************************************************************
from collections import defaultdict
from decimal import Decimal
from statistics import mean

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Batch, ReceivingNormalSKR, Sizing, FieldOfficer, Store, Region

@login_required
def download_combined_excel(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)

    # === COMMON SETUP ===
    wb = Workbook()
    ws = wb.active
    ws.title = "FO Avg & Sizing"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    pink_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    def set_cell(row, col, value, bold=False, fill=None, number_format='General'):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold)
        cell.alignment = align_center
        cell.border = thin_border
        if fill:
            cell.fill = fill
        cell.number_format = number_format
        return cell

    current_row = 1  # Start at row 1

    # === PART 1: FO AVERAGE TABLE ===
    # --- same logic as in your download_fo_average --- (use your own function here, but write into ws at current_row)

    # Collect FO averages
    from models import ReceivingNormalSKR, Sizing

    samples = ReceivingNormalSKR.objects.filter(batch=batch).select_related('field_officer__store__region')
    fo_data = defaultdict(lambda: defaultdict(list))
    fo_region = {}

    regions_to_fos_data_collection = defaultdict(lambda: defaultdict(list))

    for sample in samples:
        fo = sample.field_officer
        name = fo.name
        region = fo.store.region.name if fo.store and fo.store.region else "Unknown Region"
        fo_region[name] = region

        for attr_key, display_key in {
            'good_q': 'Good Q', 'insect_damaged': 'Insect Damaged', 'mold': 'Mold',
            'immature': 'Immature', 'weight': 'Weight', 'skr': 'SKR', 'moisture': 'Moisture'
        }.items():
            value = getattr(sample, attr_key)
            if value is not None:
                decimal_value = Decimal(str(value))
                fo_data[name][display_key].append(decimal_value)
                regions_to_fos_data_collection[region][display_key].append(decimal_value)

    individual_fo_order = ["schiks", "Beatrice Gachigi", "szczcz", "scda"]
    present_individual_fo_order = [fo for fo in individual_fo_order if fo in fo_data]

    aggregate_region_order = ["Kirinyaga", "Meru", "Mwea"]
    present_aggregate_region_order = [region for region in aggregate_region_order if region in set(fo_region.values())]

    total_columns = 1 + len(present_individual_fo_order) + len(present_aggregate_region_order)
    set_cell(current_row, 1, "FO's Average", bold=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_columns)
    current_row += 1

    # Header row
    set_cell(current_row, 1, "FO", bold=True, fill=green_fill)
    fo_col_map = {}
    current_col = 2
    for fo_name in present_individual_fo_order:
        set_cell(current_row, current_col, fo_name, bold=True, fill=green_fill)
        fo_col_map[fo_name] = current_col
        current_col += 1

    if present_aggregate_region_order:
        start = current_col
        end = current_col + len(present_aggregate_region_order) - 1
        set_cell(current_row, start, "Aggregate by region", bold=True, fill=green_fill)
        ws.merge_cells(start_row=current_row, start_column=start, end_row=current_row, end_column=end)
    current_row += 1

    # Region headers
    set_cell(current_row, 1, "Region", bold=True, fill=green_fill)
    current_col = 2
    for fo_name in present_individual_fo_order:
        region = fo_region.get(fo_name, "Unknown Region")
        set_cell(current_row, current_col, region, fill=green_fill)
        current_col += 1
    for i, region in enumerate(present_aggregate_region_order):
        fill = blue_fill if i % 2 == 0 else pink_fill
        set_cell(current_row, current_col, region, fill=fill)
        current_col += 1
    current_row += 1

    # Metric rows
    metrics = ["Good Q", "Insect Damaged", "Mold", "Immature", "Weight", "SKR", "Moisture"]
    for metric in metrics:
        set_cell(current_row, 1, metric, bold=True)
        current_col = 2
        for fo in present_individual_fo_order:
            values = fo_data[fo][metric]
            avg = round(mean(values), 2) if values else ""
            set_cell(current_row, current_col, avg, bold=True, fill=yellow_fill, number_format='0.00')
            current_col += 1

        for i, region in enumerate(present_aggregate_region_order):
            formula_cells = [
                f"{get_column_letter(fo_col_map[fo])}{current_row}"
                for fo in present_individual_fo_order
                if fo_region.get(fo) == region and fo_data[fo][metric]
            ]
            formula = f"=AVERAGE({','.join(formula_cells)})" if formula_cells else ""
            fill = blue_fill if i % 2 == 0 else pink_fill
            set_cell(current_row, current_col, formula, bold=True, fill=fill, number_format='0.00')
            current_col += 1
        current_row += 1

    # Leave 5 blank rows
    current_row += 5

    # === PART 2: SIZING TABLE ===

    sizing_data = Sizing.objects.filter(normal_skr_entry__batch=batch).select_related(
        'normal_skr_entry__field_officer').order_by('normal_skr_entry__field_officer__name')

    fo_sizing_data = defaultdict(lambda: {
        'samples': [], 'size_0_total': Decimal(0), 'size_1l_total': Decimal(0),
        'size_1s_total': Decimal(0), 'size_1xs_total': Decimal(0), 'sample_count': 0
    })
    all_sizes_total = defaultdict(Decimal)
    grand_sample_count = 0

    for entry in sizing_data:
        fo_name = entry.normal_skr_entry.field_officer.name
        sample_num = entry.normal_skr_entry.sample_number
        sample = {'sample_num': sample_num, 'size_0': entry.size_0,
                  'size_1l': entry.size_1l, 'size_1s': entry.size_1s, 'size_1xs': entry.size_1xs}
        fo_sizing_data[fo_name]['samples'].append(sample)
        fo_sizing_data[fo_name]['size_0_total'] += entry.size_0
        fo_sizing_data[fo_name]['size_1l_total'] += entry.size_1l
        fo_sizing_data[fo_name]['size_1s_total'] += entry.size_1s
        fo_sizing_data[fo_name]['size_1xs_total'] += entry.size_1xs
        fo_sizing_data[fo_name]['sample_count'] += 1
        all_sizes_total['size_0'] += entry.size_0
        all_sizes_total['size_1l'] += entry.size_1l
        all_sizes_total['size_1s'] += entry.size_1s
        all_sizes_total['size_1xs'] += entry.size_1xs
        grand_sample_count += 1

    grand_avg = {
        k: (v / grand_sample_count).quantize(Decimal('0.01')) if grand_sample_count else Decimal('0.00')
        for k, v in all_sizes_total.items()
    }

    field_officer_names = sorted(fo_sizing_data.keys())
    max_samples = max((len(data['samples']) for data in fo_sizing_data.values()), default=0)

    col = 2
    for fo_name in field_officer_names:
        ws.merge_cells(start_row=current_row, start_column=col,
                       end_row=current_row, end_column=col + max_samples - 1)
        set_cell(current_row, col, fo_name, bold=True, fill=yellow_fill)
        col += max_samples
    set_cell(current_row, col, "Average", bold=True, fill=yellow_fill)
    ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row + 1, end_column=col)
    current_row += 1

    set_cell(current_row, 1, "")
    col = 2
    for fo_name in field_officer_names:
        for s in fo_sizing_data[fo_name]['samples']:
            set_cell(current_row, col, s['sample_num'], fill=grey_fill)
            col += 1
        for _ in range(max_samples - len(fo_sizing_data[fo_name]['samples'])):
            set_cell(current_row, col, "", fill=grey_fill)
            col += 1
    set_cell(current_row, col, "", fill=grey_fill)
    current_row += 1

    size_labels = ['0', '1L', '1S', '1XS']
    size_keys = ['size_0', 'size_1l', 'size_1s', 'size_1xs']
    for i, (label, key) in enumerate(zip(size_labels, size_keys)):
        set_cell(current_row, 1, label, bold=True, fill=grey_fill)
        col = 2
        for fo in field_officer_names:
            for s in fo_sizing_data[fo]['samples']:
                set_cell(current_row, col, s[key])
                col += 1
            for _ in range(max_samples - len(fo_sizing_data[fo]['samples'])):
                set_cell(current_row, col, "")
                col += 1
        set_cell(current_row, col, grand_avg.get(key, Decimal('0.00')), bold=True, fill=yellow_fill)
        current_row += 1

    # Autosize
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # === Response ===
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Combined_Batch_{batch.batch_number}.xlsx"'
    wb.save(response)
    return response


# *******************************************************
# Processing Stage Views
@login_required
def receiving_detail(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    search_query = request.GET.get('search', '').strip()
    searched = 'search' in request.GET

    # === Normal SKR Samples ===
    normal_skr_base = ReceivingNormalSKR.objects.filter(batch=batch)
    moisture_skr_base = ReceivingSKRWithMoisture.objects.filter(batch=batch)

    if search_query:
        normal_skr_base = normal_skr_base.filter(
            Q(field_officer__name__icontains=search_query) |
            Q(field_officer__store__name__icontains=search_query) |
            Q(field_officer__store__region__name__icontains=search_query) |
            Q(sample_number__icontains=search_query)
        )

        moisture_skr_base = moisture_skr_base.filter(
            Q(field_officer__name__icontains=search_query) |
            Q(field_officer__store__name__icontains=search_query) |
            Q(field_officer__store__region__name__icontains=search_query) |
            Q(sample_number__icontains=search_query)
        )

    normal_skr = normal_skr_base.order_by(
        'field_officer__store__region__name',
        'field_officer__store__name',
        'field_officer__name',
        'sample_number'
    )
    moisture_skr = moisture_skr_base.order_by(
        'field_officer__store__region__name',
        'field_officer__store__name',
        'field_officer__name',
        'sample_number'
    )

    # === Grouping Helper Function ===
    def group_samples_by_hierarchy(samples):
        grouped_data = []
        hierarchy = {}
        for entry in samples:
            region = entry.field_officer.store.region
            store = entry.field_officer.store
            officer = entry.field_officer

            if region.name not in hierarchy:
                hierarchy[region.name] = {
                    'name': region.name,
                    'stores': {},
                    'total_samples': 0
                }
            if store.name not in hierarchy[region.name]['stores']:
                hierarchy[region.name]['stores'][store.name] = {
                    'name': store.name,
                    'officers': {},
                    'total_samples': 0
                }
            if officer.name not in hierarchy[region.name]['stores'][store.name]['officers']:
                hierarchy[region.name]['stores'][store.name]['officers'][officer.name] = {
                    'name': officer.name,
                    'samples': [],
                    'total_samples': 0
                }

            hierarchy[region.name]['stores'][store.name]['officers'][officer.name]['samples'].append(entry)
            hierarchy[region.name]['stores'][store.name]['officers'][officer.name]['total_samples'] += 1
            hierarchy[region.name]['stores'][store.name]['total_samples'] += 1
            hierarchy[region.name]['total_samples'] += 1

        for region_name in sorted(hierarchy.keys()):
            region_entry = hierarchy[region_name]
            sorted_stores = []
            for store_name in sorted(region_entry['stores'].keys()):
                store_entry = region_entry['stores'][store_name]
                sorted_officers = []
                for officer_name in sorted(store_entry['officers'].keys()):
                    sorted_officers.append(store_entry['officers'][officer_name])
                store_entry['officers'] = sorted_officers
                sorted_stores.append(store_entry)
            region_entry['stores'] = sorted_stores
            grouped_data.append(region_entry)

        return grouped_data

    grouped_skr_data = group_samples_by_hierarchy(normal_skr)
    grouped_moisture_data = group_samples_by_hierarchy(moisture_skr)

    # === FO Averages (from Normal SKR only) ===
    fo_data = defaultdict(lambda: defaultdict(list))
    fo_region = {}

    for sample in normal_skr:
        fo = sample.field_officer
        name = fo.name
        region = fo.store.region.name if fo.store and fo.store.region else "Unknown Region"
        fo_region[name] = region

        for attr_key, display_key in {
            'good_q': 'Good Q', 'insect_damaged': 'Insect Damaged', 'mold': 'Mold',
            'immature': 'Immature', 'weight': 'Weight', 'skr': 'SKR', 'moisture': 'Moisture'
        }.items():
            value = getattr(sample, attr_key)
            if value is not None:
                fo_data[name][display_key].append(Decimal(str(value)))

    fo_averages_display = defaultdict(dict)
    metrics_display_order = ["Good Q", "Insect Damaged", "Mold", "Immature", "Weight", "SKR", "Moisture"]

    for metric in metrics_display_order:
        for fo_name, metrics_dict in fo_data.items():
            values = metrics_dict.get(metric, [])
            fo_averages_display[metric][fo_name] = round(mean(values), 2) if values else ""

    # === Sizing Data Aggregation ===
    sizing_data = Sizing.objects.filter(
        normal_skr_entry__batch=batch
    ).select_related(
        'normal_skr_entry__field_officer'
    ).order_by(
        'normal_skr_entry__field_officer__name',
        'normal_skr_entry__sample_number'
    )

    fo_sizing_data = defaultdict(lambda: {
        'samples': [], 'size_0_total': Decimal(0), 'size_1l_total': Decimal(0),
        'size_1s_total': Decimal(0), 'size_1xs_total': Decimal(0), 'sample_count': 0
    })
    all_sizes_total = defaultdict(Decimal)
    grand_sample_count = 0

    for entry in sizing_data:
        fo_name = entry.normal_skr_entry.field_officer.name
        sample_num = entry.normal_skr_entry.sample_number

        sample_sizing = {
            'sample_num': sample_num,
            'size_0': entry.size_0,
            'size_1l': entry.size_1l,
            'size_1s': entry.size_1s,
            'size_1xs': entry.size_1xs,
        }
        fo_sizing_data[fo_name]['samples'].append(sample_sizing)

        fo_sizing_data[fo_name]['size_0_total'] += entry.size_0
        fo_sizing_data[fo_name]['size_1l_total'] += entry.size_1l
        fo_sizing_data[fo_name]['size_1s_total'] += entry.size_1s
        fo_sizing_data[fo_name]['size_1xs_total'] += entry.size_1xs
        fo_sizing_data[fo_name]['sample_count'] += 1

        all_sizes_total['size_0'] += entry.size_0
        all_sizes_total['size_1l'] += entry.size_1l
        all_sizes_total['size_1s'] += entry.size_1s
        all_sizes_total['size_1xs'] += entry.size_1xs
        grand_sample_count += 1

    for fo_name, data in fo_sizing_data.items():
        if data['sample_count'] > 0:
            data['averages'] = {
                'size_0': (data['size_0_total'] / data['sample_count']).quantize(Decimal('0.01')),
                'size_1l': (data['size_1l_total'] / data['sample_count']).quantize(Decimal('0.01')),
                'size_1s': (data['size_1s_total'] / data['sample_count']).quantize(Decimal('0.01')),
                'size_1xs': (data['size_1xs_total'] / data['sample_count']).quantize(Decimal('0.01')),
            }
        else:
            data['averages'] = {k: Decimal('0.00') for k in ['size_0', 'size_1l', 'size_1s', 'size_1xs']}

    grand_sizing_averages = {}
    if grand_sample_count > 0:
        for key in ['size_0', 'size_1l', 'size_1s', 'size_1xs']:
            grand_sizing_averages[key] = (all_sizes_total[key] / grand_sample_count).quantize(Decimal('0.01'))
    else:
        grand_sizing_averages = {k: Decimal('0.00') for k in ['size_0', 'size_1l', 'size_1s', 'size_1xs']}

    # === Final Context ===
    context = {
        'batch': batch,
        'search_query': search_query,
        'searched': searched,
        'normal_skr': normal_skr,
        'moisture_skr': moisture_skr,
        'grouped_skr_data': grouped_skr_data,
        'grouped_moisture_data': grouped_moisture_data,
        'metric_order': metrics_display_order + ["Spillage"],
        'transposed_metric_order': [
            "Sample No", "Field Officer", "Store", "Region", "Good Q", "Insect Damaged",
            "Mold", "Immature", "Weight (kg)", "Moisture (%)", "Spillage", "SKR", "Delete"
        ],
        'fo_averages_display': fo_averages_display,
        'fo_average_fo_names': sorted(fo_data.keys()),
        'fo_average_metrics': metrics_display_order,
        'fo_sizing_data': fo_sizing_data,
        'field_officer_names': sorted(fo_sizing_data.keys()),
        'max_sizing_samples_per_fo': max(
            len(data['samples']) for data in fo_sizing_data.values()) if fo_sizing_data else 0,
        'grand_sizing_averages': grand_sizing_averages,
        'sizing_metric_definitions': [
            ('size_0', '0'),
            ('size_1l', '1L'),
            ('size_1s', '1S'),
            ('size_1xs', '1XS'),
        ],
    }

    return render(request, 'inventory/receiving_detail.html', context)

@login_required
def nut_washing_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/nut_washing.html', {'batch': batch})

@login_required
def drying_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/drying.html', {'batch': batch})

@login_required
def drier_offloading_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/drier_offloading.html', {'batch': batch})

@login_required
def cracking_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/cracking.html', {'batch': batch})

@login_required
def inshell_sorting_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/inshell_sorting.html', {'batch': batch})

@login_required
def kernel_sorting_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/kernel_sorting.html', {'batch': batch})

@login_required
def grading_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/grading.html', {'batch': batch})

@login_required
def packaging_view(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    return render(request, 'inventory/packaging.html', {'batch': batch})

