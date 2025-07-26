from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from statistics import mean

# Import all necessary models from your .models file
from .models import ReceivingNormalSKR, ReceivingSKRWithMoisture, Sizing, ReceivingFOAverage


def render_normal_skr_section(ws, batch, start_row):
    """
    Renders the Normal SKR section in the Excel sheet.
    Leverages _render_skr_generic for common logic.
    """
    return _render_skr_generic(ws, batch, start_row, ReceivingNormalSKR, is_moisture=False)


def render_skr_with_moisture_section(ws, batch, start_row):
    """
    Renders the SKR with Moisture section in the Excel sheet.
    Leverages _render_skr_generic for common logic.
    """
    return _render_skr_generic(ws, batch, start_row, ReceivingSKRWithMoisture, is_moisture=True)


def render_sizing_section(ws, batch, start_row):
    """
    Renders the Sizing section of the Excel sheet.
    This function was specifically targeted for the `FieldError`.
    The `select_related` call is made explicit for robustness.
    """
    # Explicitly chain select_related calls to ensure correct pre-fetching.
    # This addresses the potential ambiguity that might have caused the FieldError.
    samples = Sizing.objects.filter(
        normal_skr_entry__batch=batch
    ).select_related(
        'normal_skr_entry',  # Select the ReceivingNormalSKR object
        'normal_skr_entry__field_officer',  # Select the FieldOfficer related to ReceivingNormalSKR
        'normal_skr_entry__field_officer__store',  # Select the Store related to FieldOfficer
        'normal_skr_entry__field_officer__store__region'  # Select the Region related to Store
    )

    row = start_row
    headers = ["Region", "Store", "Field Officer", "Sample", "Size 0", "Size 1L", "Size 1S", "Size 1XS"]
    for idx, h in enumerate(headers, start=2):
        ws.cell(row=row, column=idx, value=h).font = Font(bold=True)
    row += 1

    for s in samples:
        region_name = None
        store_name = None
        officer_name = None
        sample_number = None

        # Robustly access nested attributes with try-except to prevent AttributeError
        # if any part of the related chain is None (e.g., if a field officer is not assigned).
        try:
            fo_entry = s.normal_skr_entry
            if fo_entry:
                sample_number = fo_entry.sample_number
                field_officer = fo_entry.field_officer
                if field_officer:
                    officer_name = field_officer.name
                    store = field_officer.store
                    if store:
                        store_name = store.name
                        region = store.region
                        if region:
                            region_name = region.name
        except AttributeError:
            # Log the error or handle it as appropriate for your application
            print(f"Warning: Could not fully retrieve related data for Sizing sample ID: {s.id}. Skipping some fields.")
            # Continue to populate available fields, even if some are None

        ws.cell(row=row, column=2, value=region_name)
        ws.cell(row=row, column=3, value=store_name)
        ws.cell(row=row, column=4, value=officer_name)
        ws.cell(row=row, column=5, value=sample_number)
        ws.cell(row=row, column=6, value=float(s.size_0))
        ws.cell(row=row, column=7, value=float(s.size_1l))
        ws.cell(row=row, column=8, value=float(s.size_1s))
        ws.cell(row=row, column=9, value=float(s.size_1xs))
        row += 1

    return row + 2


def render_fo_average_section(ws, batch, start_row):
    """
    Renders the Field Officer Averages section of the Excel sheet.
    Includes robust attribute access.
    """
    samples = ReceivingFOAverage.objects.filter(batch=batch).select_related('field_officer__store__region')
    row = start_row
    headers = ["Region", "Store", "Field Officer", "Good Q", "Insect Damaged", "Mold", "Immature", "Weight", "SKR", "Moisture"]
    for idx, h in enumerate(headers, start=2):
        ws.cell(row=row, column=idx, value=h).font = Font(bold=True)
    row += 1

    for s in samples:
        region_name = None
        store_name = None
        officer_name = None

        try:
            field_officer = s.field_officer
            if field_officer:
                officer_name = field_officer.name
                store = field_officer.store
                if store:
                    store_name = store.name
                    region = store.region
                    if region:
                        region_name = region.name
        except AttributeError:
            print(f"Warning: Could not fully retrieve related data for ReceivingFOAverage ID: {s.id}. Skipping some fields.")

        ws.cell(row=row, column=2, value=region_name)
        ws.cell(row=row, column=3, value=store_name)
        ws.cell(row=row, column=4, value=officer_name)
        ws.cell(row=row, column=5, value=float(s.good_q))
        ws.cell(row=row, column=6, value=float(s.insect_damaged))
        ws.cell(row=row, column=7, value=float(s.mold))
        ws.cell(row=row, column=8, value=float(s.immature))
        ws.cell(row=row, column=9, value=float(s.weight))
        ws.cell(row=row, column=10, value=float(s.skr))
        ws.cell(row=row, column=11, value=float(s.moisture))
        row += 1

    return row + 2


def _render_skr_generic(ws, batch, row_offset, model, is_moisture):
    """
    Helper function to render either Normal SKR or SKR with Moisture sections.
    It applies consistent styling and calculates averages/weighted values.
    Includes robust attribute access.
    """
    samples = model.objects.filter(batch=batch).select_related(
        'field_officer__store__region' # This chain works correctly for ReceivingNormalSKR/ReceivingSKRWithMoisture
    ).order_by(
        'field_officer__store__region__name',
        'field_officer__store__name',
        'field_officer__name',
        'sample_number'
    )

    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for sample in samples:
        region = None
        store = None
        officer = None
        try:
            field_officer = sample.field_officer
            if field_officer:
                officer = field_officer.name
                store_obj = field_officer.store
                if store_obj:
                    store = store_obj.name
                    region_obj = store_obj.region
                    if region_obj:
                        region = region_obj.name
        except AttributeError:
            print(f"Warning: Could not fully retrieve related data for SKR sample ID: {sample.id}. Skipping sample.")
            continue # Skip this sample if essential related data is missing for grouping

        if region and store and officer: # Only group if all essential details are available
            grouped[region][store][officer].append(sample)
        else:
            print(f"Skipping SKR sample ID {sample.id} due to missing vital related information for grouping.")


    col_offset = 2
    data_start_col = col_offset + 1
    current_col = data_start_col
    sample_cells = []

    # Define common styles for consistency
    vibrant_green_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    shouting_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))
    base_font = Font(name='Baskerville Old Face', size=12)

    def set_cell(cell, value, fill=None, bold=False):
        """Helper to set cell value and apply common styles."""
        cell.value = value
        cell.font = Font(name=base_font.name, size=base_font.size, bold=bold)
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill

    # Section header rows (e.g., Field Officer, Stores, Region, Samples)
    set_cell(ws.cell(row=row_offset, column=col_offset), "Field Officer", bold=True)
    set_cell(ws.cell(row=row_offset + 1, column=col_offset), "Stores", bold=True)
    set_cell(ws.cell(row=row_offset + 2, column=col_offset), "Region", bold=True)
    set_cell(ws.cell(row=row_offset + 3, column=col_offset), "Samples", bold=True)

    region_ranges = defaultdict(lambda: {'start':None,'end':None})
    store_ranges = defaultdict(lambda: {'start':None,'end':None})
    officer_ranges = defaultdict(lambda: {'start':None,'end':None})

    # Build columns grouped by region/store/officer and populate headers
    for region, stores in grouped.items():
        if region_ranges[region]['start'] is None: # Set start of region span
            region_ranges[region]['start'] = current_col
        for store, officers in stores.items():
            if store_ranges[store]['start'] is None: # Set start of store span
                store_ranges[store]['start'] = current_col
            for officer, entries in officers.items():
                if officer_ranges[officer]['start'] is None: # Set start of officer span
                    officer_ranges[officer]['start'] = current_col
                # Populate header cells (Field Officer, Store, Region for each column group)
                set_cell(ws.cell(row=row_offset, column=current_col), officer, vibrant_green_fill, True)
                set_cell(ws.cell(row=row_offset+1, column=current_col), store, vibrant_green_fill, True)
                set_cell(ws.cell(row=row_offset+2, column=current_col), region, vibrant_green_fill, True)
                # Populate sample numbers for the current officer
                for sample in entries:
                    set_cell(ws.cell(row=row_offset+3, column=current_col), sample.sample_number, vibrant_green_fill, True)
                    sample_cells.append(sample) # Keep track of samples for data rows
                    current_col += 1
                officer_ranges[officer]['end'] = current_col - 1 # Set end of officer span
            store_ranges[store]['end'] = current_col - 1 # Set end of store span
        region_ranges[region]['end'] = current_col - 1 # Set end of region span

    # Merge headers for region/store/officer spans to create single, wide headers
    for group_ranges, header_row in (
        (officer_ranges, row_offset),
        (store_ranges, row_offset+1),
        (region_ranges, row_offset+2)
    ):
        for name, span in group_ranges.items():
            # Only merge if a span exists and covers more than one column
            if span['start'] is not None and span['end'] is not None and span['start'] < span['end']:
                ws.merge_cells(start_row=header_row, start_column=span['start'],
                               end_row=header_row, end_column=span['end'])

    # Define the metrics (data rows) to be displayed for each sample
    metrics = [
        ("Good Q", lambda s: s.good_q, False),
        ("Insect Damaged", lambda s: s.insect_damaged, False),
        ("Mold", lambda s: s.mold, False),
        ("Immature", lambda s: s.immature, False),
        ("Weight", lambda s: s.weight, False),
        ("Moisture", lambda s: getattr(s, 'moisture', 0.0), False), # Use 0.0 as default for moisture if not present
        ("SKR", lambda s: s.skr, True), # SKR is a calculated property or formula
        ("spillage", lambda s: getattr(s,'spillage',0.0), False), # Use 0.0 as default for spillage if not present
    ]

    metric_start = row_offset + 4 # Starting row for data metrics
    sample_count = len(sample_cells)
    avg_col = data_start_col + sample_count # Column index for Averages
    weighted_col = avg_col + 1 # Column index for Weighted Averages

    weight_row = None
    spillage_row = None
    moisture_row = None

    # Populate data for each metric row and add average formulas
    for i, (label, extractor, is_formula) in enumerate(metrics):
        r = metric_start + i # Current row for the metric
        # Set the metric label in the side column
        set_cell(ws.cell(row=r, column=col_offset), label,
                 shouting_yellow_fill if label in ["Good Q","Insect Damaged","Mold","Immature"] else None, True)

        # Keep track of row numbers for specific metrics needed in Excel formulas
        if label.lower() == "weight": weight_row = r
        if label.lower() == "spillage": spillage_row = r
        if label.lower() == "moisture": moisture_row = r

        # Populate data for each sample column
        for j, sample in enumerate(sample_cells):
            c = data_start_col + j # Current data column
            cell = ws.cell(row=r, column=c)
            # Handle SKR calculation based on whether it's a moisture batch
            if is_formula and label == "SKR":
                gq_cell = f"{get_column_letter(c)}{metric_start}" # Reference to Good Q cell in the same column
                if is_moisture and moisture_row:
                    mo_avg = get_column_letter(c) + str(moisture_row) # Reference to Moisture cell in the same column
                    # Formula for SKR with moisture content
                    formula = f"=IFERROR({gq_cell}/(100/(100-55-1.5-{mo_avg})),0)"
                else:
                    # Formula for SKR without moisture (assumes fixed divisor)
                    formula = f"=IFERROR({gq_cell}/5,0)"
                set_cell(cell, formula)
            else:
                # Extract value, convert to float (default to 0.0 if None or DecimalField is None)
                val = extractor(sample)
                set_cell(cell, float(val) if val is not None else 0.0,
                         shouting_yellow_fill if label in ["Good Q","Insect Damaged","Mold","Immature"] else None)
            cell.number_format = '0.00' # Format all numeric data as two decimal places

        # Add AVERAGE formula for the current metric row
        if sample_count > 0: # Only add average if there are samples
            start_col_letter = get_column_letter(data_start_col)
            end_col_letter = get_column_letter(data_start_col + sample_count - 1)
            avg_cell = ws.cell(row=r, column=avg_col)
            formula = f"=AVERAGE({start_col_letter}{r}:{end_col_letter}{r})"
            set_cell(avg_cell, formula,
                     shouting_yellow_fill if label in ["Good Q","Insect Damaged","Mold","Immature"]
                     else vibrant_green_fill if label=="SKR" else None)
            avg_cell.number_format = '0.00'

    # Add WEIGHTED AVERAGE formula column for relevant metrics
    for i, (label, *_rest) in enumerate(metrics):
        r = metric_start + i
        wcell = ws.cell(row=r, column=weighted_col)
        # Apply weighted average only to specific metrics
        if label not in ["Weight","Good Q","Insect Damaged","Mold","Immature","SKR","Moisture"]:
            set_cell(wcell, "N/A", bold=True)
        else:
            terms = []
            if sample_count > 0 and weight_row is not None:
                for j in range(sample_count):
                    col = get_column_letter(data_start_col + j)
                    weight_ref = f"${col}${weight_row}" # Absolute reference to weight in this column
                    metric_ref = f"{col}{r}" # Relative reference to current metric in this column
                    terms.append(f"({metric_ref}*{weight_ref})")
            numerator = "+".join(terms) if terms else "0"

            # Calculate total weight in the batch for the denominator
            total_weight_formula = "0"
            if sample_count > 0 and weight_row is not None:
                total_weight_range_start = f"${get_column_letter(data_start_col)}${weight_row}"
                total_weight_range_end = f"${get_column_letter(data_start_col+sample_count-1)}${weight_row}"
                total_weight_formula = f"SUM({total_weight_range_start}:{total_weight_range_end})"

            # Include spillage in the denominator if spillage data is available
            spv = "0"
            if spillage_row is not None:
                spv_ref = f"{get_column_letter(avg_col)}{spillage_row}" # Spillage value from average column
                spv = f"IF(ISBLANK({spv_ref}),0,{spv_ref})" # Handle blank spillage cell gracefully
            denom = f"({total_weight_formula}-{spv})"
            # Final weighted average formula, avoiding division by zero
            formula = f'=IF({denom}=0,"N/A",({numerator})/({denom}))'
            set_cell(wcell, formula,
                     shouting_yellow_fill if label in ["Good Q","Insect Damaged","Mold","Immature"]
                     else vibrant_green_fill if label=="SKR" else None)
            wcell.number_format = '0.00' # Format as two decimal places

    # Final headers for Average and Weighted columns
    set_cell(ws.cell(row=row_offset+3, column=avg_col), "Average", vibrant_green_fill, True)
    set_cell(ws.cell(row=row_offset+3, column=weighted_col), "Weighted", vibrant_green_fill, True)

    # Apply bold font and thick borders to the side column (metric labels)
    for r in range(row_offset, metric_start + len(metrics)):
        cell = ws.cell(row=r, column=col_offset)
        cell.font = Font(name=base_font.name, size=base_font.size, bold=True)
        cell.border = thick_border

    # Set appropriate column widths for better readability
    for col_idx in range(col_offset, weighted_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Return the next available row number for subsequent sections
    return metric_start + len(metrics) + 2
